"""Parse V5学长 Excel → JSON for Supabase import."""
import json
import re
import os
import sys
os.chdir(r"C:\Users\Lenovo\Desktop\chunzhao")  # Avoid inspect.py conflict
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

EXCEL_PATH = r"C:\Users\Lenovo\Desktop\历史人物资料\002\新建文件夹\2026届校招信息汇总表（微信公众号：V5学长  欢迎关注.xlsx"
OUTPUT_PATH = r"C:\Users\Lenovo\Desktop\chunzhao\scripts\jobs-data.json"

def parse_deadline_date(text):
    if not text:
        return None
    m = re.search(r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})', str(text))
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"

def extract_city(location):
    if not location:
        return None
    first = re.split(r'[\s,，、]+', str(location).strip())[0]
    return first if first else None

def clean_industry_tags(industry):
    if not industry:
        return []
    parts = re.split(r'[/／]', str(industry))
    return [p.strip() for p in parts if p.strip() and len(p.strip()) < 10][:3]

def clean_referral(raw):
    if not raw or str(raw).strip() in ('\\', ''):
        return None
    s = str(raw).strip()
    m = re.search(r'内推码[：:]?\s*(\S+)', s)
    if m:
        return m.group(1)
    if len(s) < 30 and s != '\\':
        return s
    return None

def main():
    print("📖 Loading Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=False)
    ws = wb["2026春招表"]

    # Headers at row 3 (1-indexed)
    headers = [cell.value for cell in ws[3]]
    print(f"📋 Headers: {headers}")

    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h.strip()] = i
    print(f"📊 Column map: {col_map}")

    jobs = []
    seen = set()
    link_count = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=False), start=4):
        cells = list(row)

        company = str(cells[col_map.get('公司名称', 1)].value or '').strip()
        title = str(cells[col_map.get('招聘岗位', 6)].value or '').strip()

        if not company or not title:
            continue

        key = f"{company}||{title}"
        if key in seen:
            continue
        seen.add(key)

        company_type = str(cells[col_map.get('企业性质', 3)].value or '').strip()
        industry = str(cells[col_map.get('行业大类', 4)].value or '').strip()
        target_grad = str(cells[col_map.get('招聘对象', 5)].value or '').strip()
        deadline = str(cells[col_map.get('截止时间', 10)].value or '').strip()
        city = extract_city(str(cells[col_map.get('工作地点', 8)].value or ''))
        referral = clean_referral(cells[col_map.get('内推码/备注', 13)].value)

        # Extract hyperlink from 投递方式 column
        jd_url = None
        delivery_cell = cells[col_map.get('投递方式', 12)]
        if delivery_cell.hyperlink and delivery_cell.hyperlink.target:
            jd_url = delivery_cell.hyperlink.target
            link_count += 1
        elif delivery_cell.value and str(delivery_cell.value).startswith('http'):
            jd_url = str(delivery_cell.value)
            link_count += 1

        # Build tags
        tags = []
        if company_type:
            tags.append(company_type)
        tags.extend(clean_industry_tags(industry))
        if referral:
            tags.append('有内推')
        tags = list(dict.fromkeys(tags))  # dedupe preserving order

        jobs.append({
            'title': title,
            'company': company,
            'city': city,
            'deadline': deadline if deadline else None,
            'deadline_date': parse_deadline_date(deadline),
            'tags': tags,
            'jd_url': jd_url,
            'description': None,
            'resume_tips': None,
            'evaluation': None,
            'risk_notes': None,
            'status': 'active',
            'company_type': company_type if company_type else None,
            'target_graduates': target_grad if target_grad else None,
            'referral_code': referral,
            'source': 'V5学长2026春招表',
        })

    print(f"\n📦 Parsed {len(jobs)} unique jobs")
    print(f"   Links: {link_count}")
    print(f"   With date: {sum(1 for j in jobs if j['deadline_date'])}")
    print(f"   With referral: {sum(1 for j in jobs if j['referral_code'])}")

    # Save first 500
    output = jobs[:500]
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n💾 Saved {len(output)} jobs to {OUTPUT_PATH}")

if __name__ == '__main__':
    main()
